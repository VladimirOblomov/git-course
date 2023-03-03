from random import randint
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

print('HI piple')
columns_count = int(input('Ввидите ширину маирицы: ')) 
rows_count = int(input('Ввидите высоту матрицы: '))

# создаем книгу 
wb = Workbook()
# делаем единственный лист активным 
ws = wb.active
ws.title = input("Введите название страницы: ")
for current_colum in range(1, columns_count+1):
    for current_row in range(1, rows_count+1):
        ws.cell(row=current_row, column=current_colum).value = randint(0, 101)
        
wb.save("AZAZA.xls") 
